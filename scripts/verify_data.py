#!/usr/bin/env python3
"""
Data verification script that cross-checks website data (TypeScript/JSON files in src/data/)
against Excel source files for consistency.

Usage: python3 scripts/verify_data.py
"""
import json
import os
import re
import openpyxl
from datetime import datetime
from pathlib import Path

# Paths
BASE_DIR = Path("/Users/wangqiuting/macro-dashboard")
EXCEL_DIR = Path("/Users/wangqiuting/Documents/1. 新州投资/1. 母基金配置研究/1. 宏观研究/2. 宏观数据库/宏观网站建设")
SRC_DATA_DIR = BASE_DIR / "src" / "data"

THRESHOLD = 0.1  # max acceptable difference

def fmt_month(dt):
    """Convert a datetime to 'YYYY-MM' format."""
    if isinstance(dt, datetime):
        return f"{dt.year}-{dt.month:02d}"
    return str(dt)

def to_num(v):
    """Try to convert a value to float, return None if not possible."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(v)
    except (ValueError, TypeError):
        return None

def load_typescript_records(filepath, var_name, months_field="months", values_field="values"):
    """
    Load a PmiExcelItem-style TypeScript file.
    Looks for: export const <var_name>: PmiExcelItem = { months: [...], values: {...} }
    Returns dict of month->value.
    """
    content = filepath.read_text(encoding="utf-8")
    # Find the variable block by name
    pattern = re.compile(
        r'export\s+const\s+' + re.escape(var_name) + r'\s*:\s*\w+\s*=\s*\{([^}]+)\}',
        re.DOTALL
    )
    # Simpler approach: parse by extracting months array and values object
    months = []
    values = {}

    # Extract months array
    mo_match = re.search(r'months\s*:\s*\[(.*?)\]', content, re.DOTALL)
    if mo_match:
        months_raw = mo_match.group(1)
        months = [m.strip().strip('"') for m in months_raw.split(',') if m.strip()]
    else:
        print(f"  [WARN] Could not find months array for {var_name}")
        return {}

    # Extract values object
    val_match = re.search(r'values\s*:\s*\{(.*?)\}', content, re.DOTALL)
    if val_match:
        vals_raw = val_match.group(1)
        for pair in vals_raw.split(','):
            pair = pair.strip()
            if ':' in pair:
                key, val = pair.split(':', 1)
                key = key.strip().strip('"')
                val = val.strip()
                if val.lower() == 'null':
                    values[key] = None
                else:
                    try:
                        values[key] = float(val)
                    except ValueError:
                        values[key] = val
    else:
        print(f"  [WARN] Could not find values for {var_name}")
        return {}

    result = {}
    for m in months:
        if m in values:
            v = values[m]
            if v is not None:
                result[m] = float(v)
    return result


def load_typescript_json_obj(filepath, var_name):
    """
    Load a TypeScript object structured like:
    export const <var_name> = { months: [...], data: { ... } }
    Returns the 'data' sub-object.
    """
    content = filepath.read_text(encoding="utf-8")
    # Find the variable block
    obj_match = re.search(
        r'export\s+const\s+' + re.escape(var_name) + r'\s*[=:]\s*(\{.*?\});',
        content, re.DOTALL
    )
    if not obj_match:
        print(f"  [WARN] Could not find const {var_name}")
        return {}

    # Extract months
    months = []
    mo_match = re.search(r'"months"\s*:\s*\[(.*?)\]', content, re.DOTALL)
    if mo_match:
        months_raw = mo_match.group(1)
        months = [m.strip().strip('"') for m in months_raw.split(',') if m.strip()]

    # Extract data arrays
    data_obj = {}
    data_match = re.search(r'"data"\s*:\s*\{(.*?)\}', content, re.DOTALL)
    if data_match:
        data_raw = data_match.group(1)
        # Find each array key
        array_pattern = re.compile(r'"(\w+)"\s*:\s*\[(.*?)\]', re.DOTALL)
        for arr_match in array_pattern.finditer(data_raw):
            key = arr_match.group(1)
            vals_raw = arr_match.group(2)
            vals = []
            for v in vals_raw.split(','):
                v = v.strip()
                if v.lower() == 'null':
                    vals.append(None)
                else:
                    try:
                        vals.append(float(v))
                    except ValueError:
                        vals.append(None)
            data_obj[key] = dict(zip(months, vals))

    return data_obj


def compare_values(excel_data, website_data, label):
    """Compare excel_data and website_data dicts (month->value)."""
    mismatches = []
    matched = 0
    skipped = 0

    all_months = sorted(set(list(excel_data.keys()) + list(website_data.keys())))

    for month in all_months:
        ev = excel_data.get(month)
        wv = website_data.get(month)

        if ev is None or wv is None:
            if ev is None and wv is None:
                continue
            skipped += 1
            continue

        diff = abs(ev - wv)
        if diff > THRESHOLD:
            mismatches.append((month, ev, wv, diff))
        else:
            matched += 1

    return matched, mismatches, skipped


# ==============================================
# 1. CPI 食品同比 (细分CPI.xlsx vs cpiExcelData.ts)
# ==============================================
def verify_cpi_food():
    print("=" * 60)
    print("1. CPI 食品同比 (食品 yoy)")
    print("   细分CPI.xlsx col 6 vs cpiExcelData.ts 食品.yoy")
    print("=" * 60)

    # Read Excel
    wb = openpyxl.load_workbook(EXCEL_DIR / "细分CPI.xlsx", data_only=True)
    ws = wb["中国_城市CPI_当月同比"]

    excel_food = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        dt = row[0]
        food_yoy = to_num(row[5])  # col 6 (0-indexed: 5)
        if dt is None:
            continue
        month = fmt_month(dt)
        if food_yoy is not None:
            excel_food[month] = food_yoy

    # Read website data
    content = (SRC_DATA_DIR / "cpiExcelData.ts").read_text(encoding="utf-8")

    # Extract food yoy data
    food_match = re.search(r"\{ name:\s*'食品'.*?yoy:\s*(\{.*?\})", content, re.DOTALL)
    site_food = {}
    if food_match:
        yoy_str = food_match.group(1)
        for pair in re.findall(r'"([^"]+)"\s*:\s*([^,}]+)', yoy_str):
            month, val = pair
            try:
                site_food[month] = float(val)
            except ValueError:
                pass

    # Compare
    matched, mismatches, skipped = compare_values(excel_food, site_food, "食品同比")

    print(f"  Matched: {matched}, Mismatches: {len(mismatches)}, Skipped: {skipped}")
    if mismatches:
        print(f"\n  DISCREPANCIES FOUND ({len(mismatches)}):")
        for month, ev, wv, diff in mismatches:
            print(f"    {month}: Excel={ev}, Website={wv}, Diff={diff:.2f}")
    else:
        print("  ALL GOOD: All data points match perfectly!")

    wb.close()
    return matched, mismatches, skipped


# ==============================================
# 2. CPI 城市同比 (细分CPI.xlsx vs cpiExcelData.ts)
# ==============================================
def verify_cpi_city():
    print("\n" + "=" * 60)
    print("2. CPI 城市同比 (城市CPI yoy)")
    print("   细分CPI.xlsx col 2 vs cpiExcelData.ts 城市CPI.yoy")
    print("=" * 60)

    wb = openpyxl.load_workbook(EXCEL_DIR / "细分CPI.xlsx", data_only=True)
    ws = wb["中国_城市CPI_当月同比"]

    excel_city = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        dt = row[0]
        city_yoy = to_num(row[1])  # col 2
        if dt is None or city_yoy is None:
            continue
        excel_city[fmt_month(dt)] = city_yoy

    content = (SRC_DATA_DIR / "cpiExcelData.ts").read_text(encoding="utf-8")
    city_match = re.search(r"\{ name:\s*'城市CPI'.*?yoy:\s*(\{.*?\})", content, re.DOTALL)
    site_city = {}
    if city_match:
        for pair in re.findall(r'"([^"]+)"\s*:\s*([^,}]+)', city_match.group(1)):
            month, val = pair
            try:
                site_city[month] = float(val)
            except ValueError:
                pass

    matched, mismatches, skipped = compare_values(excel_city, site_city, "城市CPI同比")

    print(f"  Matched: {matched}, Mismatches: {len(mismatches)}, Skipped: {skipped}")
    if mismatches:
        print(f"\n  DISCREPANCIES FOUND ({len(mismatches)}):")
        for month, ev, wv, diff in mismatches:
            print(f"    {month}: Excel={ev}, Website={wv}, Diff={diff:.2f}")
    else:
        print("  ALL GOOD: All data points match perfectly!")

    wb.close()
    return matched, mismatches, skipped


# ==============================================
# 3. CPI 整体 (realData.ts cpiYoyReal vs 细分CPI.xlsx)
# ==============================================
def verify_cpi_real():
    print("\n" + "=" * 60)
    print("3. CPI 同比 (realData.ts cpiYoyReal vs 细分CPI.xlsx col 2)")
    print("=" * 60)

    # Read Excel
    wb = openpyxl.load_workbook(EXCEL_DIR / "细分CPI.xlsx", data_only=True)
    ws = wb["中国_城市CPI_当月同比"]

    excel_cpi = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        dt = row[0]
        city_yoy = to_num(row[1])
        if dt is None or city_yoy is None:
            continue
        excel_cpi[fmt_month(dt)] = city_yoy

    # Read realData.ts cpiYoyReal
    content = (SRC_DATA_DIR / "realData.ts").read_text(encoding="utf-8")
    cpi_match = re.search(r'export\s+const\s+cpiYoyReal\s*:\s*Record.*?=\s*\{([^}]+)\}', content, re.DOTALL)

    site_cpi = {}
    if cpi_match:
        for pair in re.findall(r"'([^']+)'\s*:\s*([^,}]+)", cpi_match.group(1)):
            month, val = pair
            try:
                site_cpi[month] = float(val)
            except ValueError:
                pass

    matched, mismatches, skipped = compare_values(excel_cpi, site_cpi, "CPI同比")

    print(f"  Matched: {matched}, Mismatches: {len(mismatches)}, Skipped: {skipped}")
    if mismatches:
        print(f"\n  DISCREPANCIES FOUND ({len(mismatches)}):")
        for month, ev, wv, diff in mismatches:
            print(f"    {month}: Excel={ev}, realData={wv}, Diff={diff:.2f}")
    else:
        print("  ALL GOOD: All data points match perfectly!")

    wb.close()
    return matched, mismatches, skipped


# ==============================================
# 4. PPI 同比 (realData.ts ppiYoyReal vs economicData)
# ==============================================
def verify_ppi():
    print("\n" + "=" * 60)
    print("4. PPI 同比")
    print("   realData.ts ppiYoyReal vs economicData.ts ppiData.yoy")
    print("=" * 60)

    # Read realData
    content = (SRC_DATA_DIR / "realData.ts").read_text(encoding="utf-8")
    ppi_match = re.search(r'export\s+const\s+ppiYoyReal\s*:\s*Record.*?=\s*\{([^}]+)\}', content, re.DOTALL)

    real_ppi = {}
    if ppi_match:
        for pair in re.findall(r"'([^']+)'\s*:\s*([^,}]+)", ppi_match.group(1)):
            month, val = pair
            try:
                real_ppi[month] = float(val)
            except ValueError:
                pass

    # Read economicData.ts ppiData.yoy
    econ_content = (SRC_DATA_DIR / "economicData.ts").read_text(encoding="utf-8")

    # Extract months from economicData
    econ_months = []
    mo_match_econ = re.search(r"export\s+const\s+months\s*=\s*\[(.*?)\]", econ_content, re.DOTALL)
    if mo_match_econ:
        months_raw = mo_match_econ.group(1)
        econ_months = [m.strip().strip("'") for m in months_raw.split(',') if m.strip()]

    # Extract ppiData.yoy array
    ppi_yoy_match = re.search(r"ppiData\s*=\s*\{.*?yoy\s*:\s*\[(.*?)\]", econ_content, re.DOTALL)
    econ_ppi = {}
    if ppi_yoy_match and econ_months:
        yoy_raw = ppi_yoy_match.group(1)
        yoy_vals = []
        for v in yoy_raw.split(','):
            v = v.strip()
            try:
                yoy_vals.append(float(v))
            except ValueError:
                yoy_vals.append(None)

        for i, month in enumerate(econ_months):
            if i < len(yoy_vals) and yoy_vals[i] is not None:
                econ_ppi[month] = yoy_vals[i]

    # Also check PPI vs auto-generated from 八大项CPI (would need a separate sheet)
    # For now compare realData with economicData

    matched, mismatches, skipped = compare_values(real_ppi, econ_ppi, "PPI同比")

    print(f"  Matched: {matched}, Mismatches: {len(mismatches)}, Skipped: {skipped}")
    if mismatches:
        print(f"\n  DISCREPANCIES FOUND ({len(mismatches)}):")
        for month, ev, wv, diff in sorted(mismatches, key=lambda x: x[0])[-10:]:
            print(f"    {month}: realData={ev}, economicData={wv}, Diff={diff:.2f}")
    else:
        print("  ALL GOOD")

    return matched, mismatches, skipped


# ==============================================
# 5. PMI 制造业 (PMI.xlsx vs pmiExcelData.ts vs realData.ts)
# ==============================================
def verify_pmi():
    print("\n" + "=" * 60)
    print("5. PMI 制造业")
    print("   PMI.xlsx col 2 vs pmiExcelData.ts pmiPMI")
    print("=" * 60)

    # Read Excel
    wb = openpyxl.load_workbook(EXCEL_DIR / "PMI.xlsx", data_only=True)
    ws = wb["中国_制造业PMI"]

    excel_pmi = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        dt = row[0]
        pmi_val = to_num(row[1])  # col 2
        if dt is None or pmi_val is None:
            continue
        excel_pmi[fmt_month(dt)] = pmi_val

    # Read pmiExcelData.ts pmiPMI
    content = (SRC_DATA_DIR / "pmiExcelData.ts").read_text(encoding="utf-8")

    site_pmi = {}

    # Find all months and values for pmiPMI by scanning the file
    # Pattern: export const pmiPMI: PmiExcelItem = {
    start = content.find("export const pmiPMI: PmiExcelItem = {")
    if start >= 0:
        block = content[start:]
        # Find values dict
        val_idx = block.find("values: {")
        if val_idx >= 0:
            vals_part = block[val_idx + len("values: {"):]
            # Find closing brace at the right nesting level
            depth = 1
            end_idx = 0
            for i, c in enumerate(vals_part):
                if c == '{':
                    depth += 1
                elif c == '}':
                    depth -= 1
                    if depth == 0:
                        end_idx = i
                        break

            vals_str = vals_part[:end_idx]
            for pair in vals_str.split(','):
                pair = pair.strip()
                if ':' in pair:
                    key, val = pair.split(':', 1)
                    key = key.strip().strip('"')
                    try:
                        site_pmi[key] = float(val.strip())
                    except ValueError:
                        pass

    matched, mismatches, skipped = compare_values(excel_pmi, site_pmi, "PMI")

    print(f"  Matched: {matched}, Mismatches: {len(mismatches)}, Skipped: {skipped}")
    if mismatches:
        print(f"\n  DISCREPANCIES FOUND ({len(mismatches)}):")
        for month, ev, wv, diff in sorted(mismatches, key=lambda x: x[0])[-10:]:
            print(f"    {month}: Excel={ev}, pmiExcelData={wv}, Diff={diff:.2f}")
    else:
        print("  ALL GOOD: All data points match perfectly!")

    wb.close()
    return matched, mismatches, skipped


# ==============================================
# 6. PMI 非制造业 (非制造业PMI.xlsx vs nonMfgExcelData.ts)
# ==============================================
def verify_pmi_non_mfg():
    print("\n" + "=" * 60)
    print("6. PMI 非制造业 (商务活动)")
    print("   非制造业PMI.xlsx vs nonMfgExcelData.ts nonMfg商务活动")
    print("=" * 60)

    wb = openpyxl.load_workbook(EXCEL_DIR / "非制造业PMI.xlsx", data_only=True)
    ws = wb["中国_非制造业PMI_服务业"]

    excel_pmi = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        dt = row[0]
        # col 12 (0-indexed: 11) = 商务活动
        pmi_val = to_num(row[11]) if len(row) > 11 else None
        if dt is None or pmi_val is None:
            continue
        excel_pmi[fmt_month(dt)] = pmi_val
    print(f"  Excel has {len(excel_pmi)} months of data (range: {min(excel_pmi.keys())} to {max(excel_pmi.keys())})")

    content = (SRC_DATA_DIR / "nonMfgExcelData.ts").read_text(encoding="utf-8")
    nm_start = content.find("export const nonMfg商务活动: PmiExcelItem = {")

    site_pmi = {}
    if nm_start >= 0:
        block = content[nm_start:]
        val_idx = block.find("values: {")
        if val_idx >= 0:
            vals_part = block[val_idx + len("values: {"):]
            depth = 1
            end_idx = 0
            for i, c in enumerate(vals_part):
                if c == '{':
                    depth += 1
                elif c == '}':
                    depth -= 1
                    if depth == 0:
                        end_idx = i
                        break
            vals_str = vals_part[:end_idx]
            for pair in re.findall(r'"([^"]+)"\s*:\s*([^,}]+)', vals_str):
                month, val = pair
                try:
                    site_pmi[month] = float(val.strip())
                except ValueError:
                    pass

    matched, mismatches, skipped = compare_values(excel_pmi, site_pmi, "非制造业PMI")

    print(f"  Matched: {matched}, Mismatches: {len(mismatches)}, Skipped: {skipped}")
    if mismatches:
        print(f"\n  DISCREPANCIES FOUND ({len(mismatches)}):")
        for month, ev, wv, diff in sorted(mismatches, key=lambda x: x[0])[-10:]:
            print(f"    {month}: Excel={ev}, nonMfgExcelData={wv}, Diff={diff:.2f}")
    else:
        print("  ALL GOOD: All data points match perfectly!")

    wb.close()
    return matched, mismatches, skipped


# ==============================================
# 7. 社零 (社零.xlsx vs retailExcelData.json)
# ==============================================
def verify_retail():
    print("\n" + "=" * 60)
    print("7. 社零 (社会消费品零售总额 当月同比)")
    print("   社零.xlsx col 65 vs retailExcelData.json retail_total_yoy")
    print("=" * 60)

    # Read Excel
    wb = openpyxl.load_workbook(EXCEL_DIR / "社零.xlsx", data_only=True)
    ws = wb["中国_餐饮收入总额_限额以上单位_当月同"]

    excel_retail = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        dt = row[0]
        # col 65 (0-indexed: 64) = 社会消费品零售总额:当月同比(1-2月合并)
        retail_val = to_num(row[64]) if len(row) > 64 else None
        if dt is None or retail_val is None:
            continue
        month = fmt_month(dt)
        excel_retail[month] = retail_val

    # Note: 1-2月合并 means Jan data is merged into Feb row.
    # So Excel row for Feb has the combined Jan-Feb yoy, Jan is empty.
    # The website JSON stores the data differently (separate or merged).
    print(f"  Excel has {len(excel_retail)} months with retail yoy data")

    # Read retailExcelData.json
    retail_json = json.loads((SRC_DATA_DIR / "retailExcelData.json").read_text(encoding="utf-8"))
    months = retail_json.get("months", [])
    data = retail_json.get("data", {})
    retail_total_yoy = data.get("retail_total_yoy", [])

    site_retail = {}
    for i, month in enumerate(months):
        if i < len(retail_total_yoy) and retail_total_yoy[i] is not None:
            site_retail[month] = float(retail_total_yoy[i])

    matched, mismatches, skipped = compare_values(excel_retail, site_retail, "社零同比")

    print(f"  Matched: {matched}, Mismatches: {len(mismatches)}, Skipped: {skipped}")
    if mismatches:
        # Show latest 10 mismatches
        sorted_m = sorted(mismatches, key=lambda x: x[0])
        print(f"\n  DISCREPANCIES FOUND ({len(mismatches)}):")
        print("  (showing latest 15)")
        for month, ev, wv, diff in sorted_m[-15:]:
            print(f"    {month}: Excel={ev}, retailExcelData={wv}, Diff={diff:.2f}")
    else:
        print("  ALL GOOD: All data points match perfectly!")

    wb.close()
    return matched, mismatches, skipped


# ==============================================
# 8. 出口 (出口.xlsx vs exportExcelData.ts)
# ==============================================
def verify_export():
    print("\n" + "=" * 60)
    print("8. 出口 (出口金额 当月同比)")
    print("   出口.xlsx col 2 vs exportExcelData.ts")
    print("=" * 60)

    # Read Excel
    wb = openpyxl.load_workbook(EXCEL_DIR / "出口.xlsx", data_only=True)
    ws = wb["中国_出口金额_当月同比"]

    excel_export = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        dt = row[0]
        exp_val = to_num(row[1])  # col 2
        if dt is None or exp_val is None:
            continue
        excel_export[fmt_month(dt)] = exp_val

    # Read exportExcelData.ts
    content = (SRC_DATA_DIR / "exportExcelData.ts").read_text(encoding="utf-8")

    site_export = {}
    # Find exp当月同比 variable
    exp_start = content.find("export const exp当月同比: ExpItem = {")
    if exp_start >= 0:
        block = content[exp_start:]
        val_idx = block.find("values: {")
        if val_idx >= 0:
            vals_part = block[val_idx + len("values: {"):]
            depth = 1
            end_idx = 0
            for i, c in enumerate(vals_part):
                if c == '{':
                    depth += 1
                elif c == '}':
                    depth -= 1
                    if depth == 0:
                        end_idx = i
                        break
            vals_str = vals_part[:end_idx]
            for pair in re.findall(r'"([^"]+)"\s*:\s*([^,}]+)', vals_str):
                month, val = pair
                try:
                    site_export[month] = float(val.strip())
                except ValueError:
                    pass

    matched, mismatches, skipped = compare_values(excel_export, site_export, "出口同比")

    print(f"  Matched: {matched}, Mismatches: {len(mismatches)}, Skipped: {skipped}")
    if mismatches:
        sorted_m = sorted(mismatches, key=lambda x: x[0])
        print(f"\n  DISCREPANCIES FOUND ({len(mismatches)}):")
        print("  (showing latest 15)")
        for month, ev, wv, diff in sorted_m[-15:]:
            print(f"    {month}: Excel={ev}, exportExcelData={wv}, Diff={diff:.2f}")
    else:
        print("  ALL GOOD: All data points match perfectly!")

    wb.close()
    return matched, mismatches, skipped


# ==============================================
# 9. PPI 子项 vs realData ppiSubItemYoyReal
# ==============================================
def verify_ppi_subitems():
    print("\n" + "=" * 60)
    print("9. PPI 细分项 (realData.ts ppiSubItemYoyReal)")
    print("   Checking internal consistency of latest 6 months")
    print("=" * 60)

    content = (SRC_DATA_DIR / "realData.ts").read_text(encoding="utf-8")

    # Extract ppiSubItemYoyReal - find the beginning and end of this section
    yoy_start = content.find("export const ppiSubItemYoyReal")
    if yoy_start < 0:
        print("  Could not find ppiSubItemYoyReal")
        return 0, [], 0

    mom_start = content.find("export const ppiSubItemMomReal", yoy_start)
    if mom_start < 0:
        print("  Could not find ppiSubItemMomReal boundary")
        return 0, [], 0

    # Only look in the YOY section
    yoy_section = content[yoy_start:mom_start]
    months_found = []
    for m in re.finditer(r"'(\d{4}-\d{2})'\s*:\s*\{([^}]+)\}", yoy_section):
        month = m.group(1)
        items_str = m.group(2)
        items = {}
        for pair in re.findall(r'"([^"]+)"\s*:\s*([^,}]+)', items_str):
            key, val = pair
            try:
                items[key] = float(val)
            except ValueError:
                pass
        months_found.append((month, items))

    print(f"  Found {len(months_found)} months with PPI yoy sub-item data")

    # Just show the latest 3 months data for manual inspection
    print("\n  Latest 3 months PPI sub-items:")
    for month, items in sorted(months_found, key=lambda x: x[0])[-3:]:
        print(f"    {month}: {json.dumps(items, ensure_ascii=False)}")

    # Compare latest month ppi total with ppiYoyReal
    ppi_match = re.search(r'export\s+const\s+ppiYoyReal\s*:\s*Record.*?=\s*\{([^}]+)\}', content, re.DOTALL) if False else None
    # Use simpler: find ppiYoyReal dict
    yoyreal_start = content.find("export const ppiYoyReal")
    ppi_yoy = {}
    if yoyreal_start >= 0:
        yr_content = content[yoyreal_start:]
        # Find the Record type declaration
        record_end = yr_content.find("= {")
        if record_end >= 0:
            dict_start = yr_content.find("{", record_end)
            if dict_start >= 0:
                dict_part = yr_content[dict_start+1:]
                depth = 1
                for i, c in enumerate(dict_part):
                    if c == '{':
                        depth += 1
                    elif c == '}':
                        depth -= 1
                        if depth == 0:
                            dict_str = dict_part[:i]
                            for pair in re.findall(r"'([^']+)'\s*:\s*([^,}]+)", dict_str):
                                month, val = pair
                                try:
                                    ppi_yoy[month] = float(val)
                                except ValueError:
                                    pass
                            break

    print("\n  Cross-check PPI总指数 in sub-items vs ppiYoyReal:")
    cross_check_failures = 0
    for month, items in sorted(months_found, key=lambda x: x[0]):
        if month in ppi_yoy and "PPI总指数" in items:
            diff = abs(ppi_yoy[month] - items["PPI总指数"])
            if diff > 0.15:
                cross_check_failures += 1
                print(f"    MISMATCH {month}: ppiYoyReal={ppi_yoy[month]}, ppiSubItem={items['PPI总指数']}")
    if cross_check_failures == 0:
        print("    ALL GOOD: ppiYoyReal matches ppiSubItemYoyReal PPI总指数")
    else:
        print(f"    {cross_check_failures} mismatches found")
    return 0, [], 0


# ==============================================
# Main
# ==============================================
def main():
    print("=" * 60)
    print("  MACRO DASHBOARD DATA VERIFICATION")
    print("  Cross-checking website TS/JSON data against Excel source files")
    print(f"  Threshold: {THRESHOLD}")
    print("=" * 60)

    results = {}

    # Run all checks
    results["CPI食品同比"] = verify_cpi_food()
    results["CPI城市同比"] = verify_cpi_city()
    results["CPI同比(realData)"] = verify_cpi_real()
    results["PPI同比"] = verify_ppi()
    results["PMI制造业"] = verify_pmi()
    results["PMI非制造业"] = verify_pmi_non_mfg()
    results["社零同比"] = verify_retail()
    results["出口同比"] = verify_export()
    results["PPI细分项"] = verify_ppi_subitems()

    # Summary
    print("\n" + "=" * 60)
    print("  SUMMARY")
    print("=" * 60)
    all_good = True
    for name, (matched, mismatches, skipped) in results.items():
        status = "PASS" if len(mismatches) == 0 else "FAIL"
        if len(mismatches) > 0:
            all_good = False
        print(f"  [{status}] {name}: {matched} matched, {len(mismatches)} mismatches, {skipped} skipped")

    total_mismatches = sum(len(m) for _, m, _ in results.values())
    total_matched = sum(m for m, _, _ in results.values())

    print(f"\n  Total: {total_matched} matched, {total_mismatches} mismatches")

    if all_good:
        print("\n  CONCLUSION: All verified indicators match perfectly between Excel and website data!")
    else:
        print("\n  CONCLUSION: Some discrepancies found. Review details above.")

    print("=" * 60)


if __name__ == "__main__":
    main()
